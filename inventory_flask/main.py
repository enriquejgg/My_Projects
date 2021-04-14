from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_sqlalchemy import SQLAlchemy
import os
import openpyxl
import pandas as pd
from openpyxl.chart import BarChart
from functools import wraps
from sys import exit, stderr, stdout
from traceback import print_exc

# la funcion que sigue la tuve que añadir pues en un momento dado PyCharm dejo de funcionar y no ejecutaba ningun codigo
# el error que mostraba era siempre este "BrokenPipeError: [Errno 32] Broken pipe"
# esta soluccion esta tomada de este enlace:
# https://gist.github.com/noahp/8fc6e5239a6358ce43bd3b8af19bacba

def suppress_broken_pipe_msg(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except SystemExit:
            raise
        except:
            print_exc()
            exit(1)
        finally:
            try:
                stdout.flush()
            finally:
                try:
                    stdout.close()
                finally:
                    try:
                        stderr.flush()
                    finally:
                        stderr.close()

    return wrapper


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database/inventory.db'
db = SQLAlchemy(app)

# conexion a la tabla Products de la base de datos

class Products(db.Model):

    __tablename__ = "products"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String)
    description = db.Column(db.String)
    stock = db.Column(db.Integer)
    price = db.Column(db.Integer)
    top_stock = db.Column(db.Integer)
    location = db.Column(db.String)

###############

###############

db.create_all()
db.session.commit()

# abrimos la homepage
# Se carga el template index.html

@app.route('/')
def home():
    return render_template("index.html")

# acceso a la pagina clients donde vemos el listado de productos

@app.route('/display_c', methods=['GET'])
def show_c():
    all_prods = Products.query.all()
    return render_template("clients.html", prod_list=all_prods)

# pagina de acceso a la parte de administracion de stock. abrimos el template para  introducir las credenciales de acceso

@app.route('/flash')
def home1():
    if not session.get('logged_in'):
        return render_template('login.html')

# verificacion de las credenciales de acceso

@app.route('/login', methods=['POST', 'GET'])
def do_admin_login():

    if request.form['password'] == 'password' and request.form['username'] == 'admin':

        session['logged_in'] = True
        return render_template("welcome_admin.html")

    else: #si las credenciales introducidas no son correctas
        flash('wrong username or password!')
        return render_template("login.html")

#######

@app.route('/admin_c', methods=['GET'])
def admin_c():
    all_prods = Products.query.all()
    stock_a = Products.query.with_entities(Products.stock).all()  # extraemos valores de las columnas
    stock_b = Products.query.with_entities(Products.top_stock).all()
    stock_a = [stock[0] for stock in
               stock_a]  # transformamos los valores de tipo class 'sqlalchemy.util._collections.result'
    stock_b = [top_stock[0] for top_stock in stock_b]  # a tipo int para poder trabajar con ellos
    ratios = []
    for i in range(len(stock_a)): #creamos un loop para ver si el stock es menos del 10% del stock maximo
        ratio = stock_a[i] / stock_b[i]
        ratios.append(ratio)

    error = 'We must order this product!!!' #si el nivel de stock es menos del 10% desplegamos este error
    return render_template("admin_c.html", prod_list_ratios=zip(all_prods, ratios), error=error)

#######
# Funcion para crear productos nuevos

@app.route('/add_c', methods=['POST'])

def add_c():

    if request.method == 'POST':
        if request.form['name'] == '' or request.form['description'] == '': #campos minimos para poder crear un producto
            flash("Name and Description fields cannot be empty")
            return redirect('admin_c')

        else:
            product = Products(name=request.form['name'], description=request.form['description'], price=request.form['price'],
                               stock=request.form['stock'], top_stock=request.form['top_stock'], location=request.form['location'])
            db.session.add(product)
            db.session.commit()
            flash('Product Added Successfully')
            return redirect(url_for('admin_c'))


#######
# Funcion para borrar productos de la DB

@app.route('/delete_c/<id>')

def delete_c(id):

    product = Products.query.filter_by(id=int(id)).delete()
    db.session.commit()
    flash('Product Deleted Successfully')
    return redirect(url_for('admin_c'))


#######
# Funcion para abrir el template donde podemos editar productos de la DB

@app.route('/edit_c/<id>', methods=['GET'])

def edit_c(id): #en el template podemos ver los valores previos del producto, antes de modificarlo

    product = Products.query.filter_by(id=int(id)).first()
    db.session.commit()
    return render_template('edit_c.html', product=product)


#######
# Funcion donde editamos el producto seleccionado

@app.route('/update_c/<int:id>', methods=['GET', 'POST'])

def update_c(id):

    product = Products.query.get_or_404(id)

    if request.method == 'POST':
        flash("Product updated Successfully")
        product.name = request.form['name_edit']
        product.description = request.form['description_edit']
        product.price = request.form['price_edit']
        product.stock = request.form['stock_edit']
        product.top_stock = request.form['top_stock_edit']
        product.location = request.form['location_edit']
        db.session.commit()
        return redirect(url_for('admin_c'))


########

# para volver a la pagina de inicio desde el area admin

@app.route('/logout')
def logout():
    session['logged_in'] = False
    return home()

# conexion a la table Suppliers

class Suppliers(db.Model):

    __tablename__ = "suppliers"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String)
    address = db.Column(db.String)
    phone = db.Column(db.Integer)
    VAT = db.Column(db.Integer)


db.create_all()
db.session.commit()

# acceso a la pagina general para ver la lista de proveedores

@app.route('/display_s', methods=['GET'])
def show_s():
    all_details = Suppliers.query.all()
    return render_template("suppliers.html", supp_list=all_details)

# acceso a la pagina de administracion donde podemos modificar los datos de estos o añadir nuevos datos

@app.route('/admin_s', methods=['GET'])
def admin_s():
    all_prods_s = Suppliers.query.all()
    return render_template("admin_s.html", supp_list_s=all_prods_s)

# funcion para crear los graficos con las cifras de compra anuales a cada proveedor

@app.route('/stats1')
def home2():

    wb1 = openpyxl.load_workbook("graph.xlsx")
    sheet_edit1 = wb1["ACME Products"]

    data = openpyxl.chart.Reference(sheet_edit1, min_col=2, min_row=2, max_col=2, max_row=12)
    categs = openpyxl.chart.Reference(sheet_edit1, min_col=1, min_row=2, max_row=12)

    graph1 = openpyxl.chart.BarChart()

    graph1.add_data(data=data)
    graph1.set_categories(categs)
    sheet_edit1.add_chart(graph1, "D3")

    graph1.title = "Purchases to ACME Products"
    graph1.varyColors = True
    graph1.y_axis.majorGridlines = None  # esto borra las lineas horizontales en el grafico

    graph1.width = 18
    graph1.height = 12

    wb1.save("graph.xlsx")

##############

    wb2 = openpyxl.load_workbook("graph.xlsx")
    sheet_edit2 = wb2["Pink Panther Supplies"]

    data = openpyxl.chart.Reference(sheet_edit2, min_col=2, min_row=2, max_col=2, max_row=12)
    categs = openpyxl.chart.Reference(sheet_edit2, min_col=1, min_row=2, max_row=12)

    graph2 = openpyxl.chart.BarChart()

    graph2.add_data(data=data)
    graph2.set_categories(categs)
    sheet_edit2.add_chart(graph2, "D3")

    graph2.title = "Purchases to Pink Panther Supplies"
    graph2.varyColors = True
    graph2.y_axis.majorGridlines = None  # esto borra las lineas horizontales en el grafico

    graph2.width = 18
    graph2.height = 12

    wb2.save("graph.xlsx")

##############

    wb3 = openpyxl.load_workbook("graph.xlsx")
    sheet_edit3 = wb3["Mortadelo y Filemon Tech"]

    data = openpyxl.chart.Reference(sheet_edit3, min_col=2, min_row=2, max_col=2, max_row=12)
    categs = openpyxl.chart.Reference(sheet_edit3, min_col=1, min_row=2, max_row=12)

    graph3 = openpyxl.chart.BarChart()

    graph3.add_data(data=data)
    graph3.set_categories(categs)
    sheet_edit3.add_chart(graph3, "D3")

    graph3.title = "Purchases to Mortadelo y Filemon Tech"
    graph3.varyColors = True
    graph3.y_axis.majorGridlines = None  # esto borra las lineas horizontales en el grafico

    graph3.width = 18
    graph3.height = 12

    wb3.save("graph.xlsx")

##############

    wb4 = openpyxl.load_workbook("graph.xlsx")
    sheet_edit4 = wb4["Donald Duck Software"]

    data = openpyxl.chart.Reference(sheet_edit4, min_col=2, min_row=2, max_col=2, max_row=12)
    categs = openpyxl.chart.Reference(sheet_edit4, min_col=1, min_row=2, max_row=12)

    graph4 = openpyxl.chart.BarChart()

    graph4.add_data(data=data)
    graph4.set_categories(categs)
    sheet_edit4.add_chart(graph4, "D3")

    graph4.title = "Purchases to Donald Duck Software"
    graph4.varyColors = True
    graph4.y_axis.majorGridlines = None  # esto borra las lineas horizontales en el grafico

    graph4.width = 18
    graph4.height = 12

    wb4.save("graph.xlsx")

##############

    wb5 = openpyxl.load_workbook("graph.xlsx")
    sheet_edit5 = wb5["Wile E Engineering"]

    data = openpyxl.chart.Reference(sheet_edit5, min_col=2, min_row=2, max_col=2, max_row=12)
    categs = openpyxl.chart.Reference(sheet_edit5, min_col=1, min_row=2, max_row=12)

    graph5 = openpyxl.chart.BarChart()

    graph5.add_data(data=data)
    graph5.set_categories(categs)
    sheet_edit5.add_chart(graph5, "D3")

    graph5.title = "Purchases to Wile E Engineering"
    graph5.varyColors = True
    graph5.y_axis.majorGridlines = None  # esto borra las lineas horizontales en el grafico

    graph5.width = 18
    graph5.height = 12

    wb5.save("graph.xlsx")

##############

    wb6 = openpyxl.load_workbook("graph.xlsx")
    sheet_edit6 = wb6["Corona Antivirus"]

    data = openpyxl.chart.Reference(sheet_edit6, min_col=2, min_row=2, max_col=2, max_row=12)
    categs = openpyxl.chart.Reference(sheet_edit6, min_col=1, min_row=2, max_row=12)

    graph6 = openpyxl.chart.BarChart()

    graph6.add_data(data=data)
    graph6.set_categories(categs)
    sheet_edit6.add_chart(graph6, "D3")

    graph6.title = "Purchases to Corona Antivirus"
    graph6.varyColors = True
    graph6.y_axis.majorGridlines = None  # esto borra las lineas horizontales en el grafico

    graph6.width = 18
    graph6.height = 12

    wb6.save("graph.xlsx")

##############

    wb7 = openpyxl.load_workbook("graph.xlsx")
    sheet_edit7 = wb7["Pepe Gotera y Otilio - Computer"]

    data = openpyxl.chart.Reference(sheet_edit7, min_col=2, min_row=2, max_col=2, max_row=12)
    categs = openpyxl.chart.Reference(sheet_edit7, min_col=1, min_row=2, max_row=12)

    graph7 = openpyxl.chart.BarChart()

    graph7.add_data(data=data)
    graph7.set_categories(categs)
    sheet_edit7.add_chart(graph7, "D3")

    graph7.title = "Purchases to Pepe Gotera y Otilio - Computer Repairs"
    graph7.varyColors = True
    graph7.y_axis.majorGridlines = None  # esto borra las lineas horizontales en el grafico

    graph7.width = 18
    graph7.height = 12

    wb7.save("graph.xlsx")

##############
    db.session.commit()
    return redirect(url_for('show_s'))

##############

# con esta funcion podemos ver los datos comparados de compras y ventas de los ultimos 10 años y ver si hubo perdidas o beneficios

@app.route('/admin_f', methods=['GET'])

def admin_f():

    #sales = pd.ExcelFile('graph.xlsx', engine='openpyxl')
    #sales.sheet_names

    financial = pd.read_excel('graph.xlsx', sheet_name='Profits And Losses', engine='openpyxl')
    #acceso a los dataframes con los datos del workbook correspondiente
    financial.set_index("Financials", inplace=True) #definimos un valor indice de la primera fila del workbook
    pl = financial.loc[['Total Sales','Total Purchases', 'P&L (Profits and Losses)']] #accedemos a las filas que queremos desplegar
    return render_template('admin_f.html', pl=pl.to_html(index=True, header=True, border=2, index_names=False, col_space=100, justify='right'))
    # pasamos los valores al template html, y modificamos los parametros para desplegarlo de una forma mas visual

##############
# Funcion para mostrar los datos de venta correspondientes a cada proveedor

@app.route('/display_indiv/<id>', methods=['GET'])

def yearly(id):

    supplier = Suppliers.query.filter_by(id=int(id)).first()
    db.session.commit()
    id_n = int(id) # para poder usar id en jinja2 tenemos que convertir (id) a int, pues la variable (id) es in string

    sales = pd.ExcelFile('graph.xlsx', engine='openpyxl')
    sales.sheet_names

    list_sales = []
    for k in range(len(sales.sheet_names) - 1): # hacemos un loop a traves de los diferentes workbooks, excepto el ultimo que es el de las ventas
        sales_supplier = pd.read_excel('graph.xlsx', sheet_name=k, engine='openpyxl', nrows=11, usecols='A:B')
        # seleccionamos que filas y columnas queremos usar
        list_sales.append(sales_supplier)
    one_sales = list_sales[id_n-1] # asi se desplegaran los datos correspondientes a cada proveedor
                                   # los valores id comienzan en 1 (se tomas de la base de datos
                                   # los valores de la lista de ventas comienzan en 0, los hacemos coincidir para
                                   # desplegar la informacion correcta

    one_sales.insert(1,'','===>') #asi podemos añadir una columna intermedia para separar los datos y hacerlos mas legibles

    return render_template('supp_sales.html', supplier_l = supplier, id_n=id_n, one_sales=one_sales.to_html
    (index=False, header=False, border=0))


#########
#Funcion para borrar proveedores de la DB

@app.route('/delete_s/<id>')

def delete_s(id):

    supplier = Suppliers.query.filter_by(id=int(id)).delete()
    db.session.commit()
    flash('Supplier Deleted Successfully')
    return redirect(url_for('admin_s'))

#######
# Funcion para crear productos nuevos

@app.route('/add_s', methods=['POST'])

def add_s():

    if request.method == 'POST':
        if request.form['name'] == '' or request.form['address'] == '' or request.form['phone_number'] == '' or request.form['VAT_number'] == '':
            flash("Fields cannot be empty") #aqui se requieren todos los campos
            return redirect('admin_s')

        else:
            supplier = Suppliers(name=request.form['name'], address=request.form['address'], phone=request.form['phone_number'],
                               VAT=request.form['VAT_number'])
            db.session.add(supplier)
            db.session.commit()
            flash('Supplier Added Successfully') #mesaja a desplegar cuando el proveedor ya esta creado
            return redirect(url_for('admin_s'))


#######
# Funcion para abrir el template donde podemos editar productos de la DB

@app.route('/edit_s/<id>', methods=['GET'])

def edit_s(id):

    supplier = Suppliers.query.filter_by(id=int(id)).first()
    db.session.commit()
    return render_template('edit_s.html', supplier=supplier)


#######
# Funcion donde editamos el producto seleccionado

@app.route('/update_s/<int:id>', methods=['GET', 'POST'])

def update_s(id):

    supplier = Suppliers.query.get_or_404(id)

    if request.method == 'POST':
        flash("Supplier Updated Successfully")
        supplier.name = request.form['name_edit']
        supplier.address = request.form['address_edit']
        supplier.phone = request.form['phone_edit']
        supplier.VAT = request.form['VAT_edit']
        db.session.commit()
        return redirect(url_for('admin_s'))


########


if __name__ == '__main__':
    app.secret_key = os.urandom(12)
    app.run(debug=True)
